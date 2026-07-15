"""
seed_bahan_baku.py
==================
Import data TKPI (Tabel Komposisi Pangan Indonesia 2020) dari file Excel
ke tabel bahan_baku di MySQL.

Sumber: Kandungan Nilai Gizi Menu 12-17 Jan 2026.xlsx → Sheet "TKPI Kemenkes 2020"

Kolom Excel → Kolom DB:
  Col 3  (NAMA BAHAN) → nama_bahan
  Col 4  (SUMBER)     → nama_baku
  Col 6  (ENERGI)     → energi
  Col 7  (PROTEIN)    → protein
  Col 8  (LEMAK)      → lemak
  Col 9  (KH)         → karbohidrat
  Col 10 (SERAT)      → serat
  "per 100 gram"      → takaran
"""

import openpyxl
from database import get_db

EXCEL_PATH = r"Kandungan Nilai Gizi Menu 12-17 Jan 2026.xlsx"
SHEET_NAME = "TKPI Kemenkes 2020"
DATA_START_ROW = 8  # Baris pertama data (setelah header)


def safe_float(val):
    """Konversi nilai ke float, return 0.0 jika gagal."""
    if val is None or val == '' or val == '-':
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def seed():
    # ── Baca Excel ──────────────────────────────────────────────────────────
    print(f"[SEED] Membaca {EXCEL_PATH} -> Sheet '{SHEET_NAME}' ...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    rows_to_insert = []
    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        nama_bahan = ws.cell(row_num, 3).value   # Col C = NAMA BAHAN
        if not nama_bahan or str(nama_bahan).strip() == '':
            continue  # Skip baris kosong

        nama_bahan = str(nama_bahan).strip()
        sumber     = str(ws.cell(row_num, 4).value or '').strip()  # Col D = SUMBER
        energi     = safe_float(ws.cell(row_num, 6).value)         # Col F = ENERGI
        protein    = safe_float(ws.cell(row_num, 7).value)         # Col G = PROTEIN
        lemak      = safe_float(ws.cell(row_num, 8).value)         # Col H = LEMAK
        karbohidrat = safe_float(ws.cell(row_num, 9).value)        # Col I = KH
        serat      = safe_float(ws.cell(row_num, 10).value)        # Col J = SERAT

        rows_to_insert.append((
            nama_bahan,
            'per 100 gram',   # takaran standar TKPI
            sumber,           # nama_baku = sumber referensi
            energi,
            karbohidrat,
            protein,
            lemak,
            serat,
        ))

    print(f"[SEED] Ditemukan {len(rows_to_insert)} bahan pangan dari TKPI.")

    if not rows_to_insert:
        print("[SEED] Tidak ada data untuk diimport. Selesai.")
        return

    # ── Insert ke MySQL ─────────────────────────────────────────────────────
    conn = get_db()
    cur  = conn.cursor()

    try:
        # Cek data yang sudah ada untuk menghindari duplikat
        cur.execute("SELECT nama_bahan FROM bahan_baku")
        existing = {row['nama_bahan'].lower() for row in cur.fetchall()}
        print(f"[SEED] Data bahan baku yang sudah ada: {len(existing)} record.")

        new_rows = [r for r in rows_to_insert if r[0].lower() not in existing]
        skipped  = len(rows_to_insert) - len(new_rows)

        if skipped > 0:
            print(f"[SEED] Melewati {skipped} bahan yang sudah ada di database.")

        if not new_rows:
            print("[SEED] Semua data sudah ada di database. Tidak ada yang ditambahkan.")
            return

        sql = """
            INSERT INTO bahan_baku
                (nama_bahan, takaran, nama_baku, energi, karbohidrat, protein, lemak, serat)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        cur.executemany(sql, new_rows)
        conn.commit()

        print(f"[SEED] OK! Berhasil memasukkan {len(new_rows)} bahan baku dari TKPI!")

    except Exception as e:
        conn.rollback()
        print(f"[SEED] ERROR: {e}")
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == '__main__':
    seed()
